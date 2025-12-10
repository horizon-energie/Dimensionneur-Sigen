from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def get_catalog():
    # ----------------------------------------------------
    # PANNEAUX
    # ----------------------------------------------------
    panels = [
        ["Trina450", 450, 52.9, 44.6, 10.74, 10.09, -0.24],
        ["Trina500", 500, 40.1, 38.3, 15.03, 12.18, -0.24],
        ["Trina505", 505, 51.7, 43.7, 12.13, 11.56, -0.25],
        ["Solux415", 415, 37.95, 31.83, 13.77, 13.04, -0.28],
        ["Solux420", 420, 38.14, 32.02, 13.85, 13.12, -0.28],
        ["Solux425", 425, 38.32, 32.20, 13.93, 13.20, -0.28],
    ]

    # ----------------------------------------------------
    # SIGENERGY — ONDULEURS COMPLETS CORRIGÉS
    # Format :
    # (ID, P_AC_nom, P_DC_max, V_MPP_min, V_MPP_max, V_DC_max,
    #  I_MPPT, Nb_MPPT, Type_reseau, Famille, V_nom_dc)
    # ----------------------------------------------------

    inverters = [
        # --- MONO Hybride ---
        ("Hybride2.0Mono", 2000, 4000, 50, 550, 600, 16, 2, "Mono", "Hybride", 350),
        ("Hybride3.0Mono", 3000, 6000, 50, 550, 600, 16, 2, "Mono", "Hybride", 350),
        ("Hybride6.0Mono", 6000, 12000, 50, 550, 600, 16, 2, "Mono", "Hybride", 350),

        # --- MONO Store ---
        ("Store3.0Mono", 3000, 6000, 50, 550, 600, 16, 2, "Mono", "Store", 350),
        ("Store3.6Mono", 3680, 7360, 50, 550, 600, 16, 2, "Mono", "Store", 350),
        ("Store4.0Mono", 4000, 8000, 50, 550, 600, 16, 2, "Mono", "Store", 350),
        ("Store4.6Mono", 4600, 9200, 50, 550, 600, 16, 2, "Mono", "Store", 350),
        ("Store6.0Mono", 6000, 12000, 50, 550, 600, 16, 2, "Mono", "Store", 350),
        ("Store8.0Mono", 8000, 16000, 50, 550, 600, 16, 3, "Mono", "Store", 350),
        ("Store10.0Mono", 10000, 20000, 50, 550, 600, 16, 4, "Mono", "Store", 350),
        ("Store12.0Mono", 12000, 24000, 50, 550, 600, 16, 4, "Mono", "Store", 350),

        # --- TRI 3x230 Hybride ---
        ("Hybride3.0Delta", 3000, 6000, 50, 550, 600, 16, 2, "Tri 3x230", "Hybride", 360),
        ("Hybride5.0Delta", 5000, 10000, 50, 550, 600, 16, 2, "Tri 3x230", "Hybride", 360),
        ("Hybride6.0Delta", 6000, 12000, 50, 550, 600, 16, 3, "Tri 3x230", "Hybride", 360),
        ("Hybride8.0Delta", 8000, 16000, 50, 550, 600, 16, 3, "Tri 3x230", "Hybride", 360),
        ("Hybride10.0Delta", 10000, 20000, 50, 550, 600, 16, 4, "Tri 3x230", "Hybride", 360),

        # --- TRI 3x230 Store ---
        ("Store6.0Delta", 6000, 12000, 50, 550, 600, 16, 2, "Tri 3x230", "Store", 360),
        ("Store8.0Delta", 8000, 16000, 50, 550, 600, 16, 3, "Tri 3x230", "Store", 360),

        # --- TRI 3x400 Hybride (Haute tension) ---
        ("Hybride3.0Tetra", 3000, 6000, 160, 1000, 1100, 16, 2, "Tri 3x400", "Hybride", 600),
        ("Hybride5.0Tetra", 5000, 10000, 160, 1000, 1100, 16, 2, "Tri 3x400", "Hybride", 600),
        ("Hybride6.0Tetra", 6000, 12000, 160, 1000, 1100, 16, 3, "Tri 3x400", "Hybride", 600),
        ("Hybride8.0Tetra", 8000, 16000, 160, 1000, 1100, 16, 3, "Tri 3x400", "Hybride", 600),
        ("Hybride10.0Tetra", 10000, 20000, 160, 1000, 1100, 16, 4, "Tri 3x400", "Hybride", 600),
        ("Hybride12.0Tetra", 12000, 24000, 160, 1000, 1100, 16, 4, "Tri 3x400", "Hybride", 600),
        ("Hybride15.0Tetra", 15000, 30000, 160, 1000, 1100, 16, 4, "Tri 3x400", "Hybride", 600),

        # --- TRI 3x400 Store (Haute tension) ---
        ("Store5.0Tetra", 5000, 10000, 160, 1000, 1100, 16, 2, "Tri 3x400", "Store", 600),
        ("Store6.0Tetra", 6000, 12000, 160, 1000, 1100, 16, 2, "Tri 3x400", "Store", 600),
        ("Store8.0Tetra", 8000, 16000, 160, 1000, 1100, 16, 3, "Tri 3x400", "Store", 600),
        ("Store10.0Tetra", 10000, 20000, 160, 1000, 1100, 16, 4, "Tri 3x400", "Store", 600),
        ("Store15.0Tetra", 15000, 30000, 160, 1000, 1100, 16, 4, "Tri 3x400", "Store", 600),
        ("Store17.0Tetra", 17000, 34000, 160, 1000, 1100, 16, 4, "Tri 3x400", "Store", 600),
        ("Store20.0Tetra", 20000, 40000, 160, 1000, 1100, 16, 4, "Tri 3x400", "Store", 600),
        ("Store25.0Tetra", 25000, 50000, 160, 1000, 1100, 16, 4, "Tri 3x400", "Store", 600),
        ("Store30.0Tetra", 30000, 60000, 160, 1000, 1100, 16, 4, "Tri 3x400", "Store", 600),
    ]

    # ----------------------------------------------------
    # BATTERIES
    # ----------------------------------------------------
    batteries = [
        ["Sigen6", 6],
        ["Sigen10", 10],
    ]

    return panels, inverters, batteries


def _autofit(ws, width=16, max_col=20):
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def generate_workbook_bytes(config: dict) -> bytes:
    panels, inverters, batteries = get_catalog()
    wb = Workbook()

    # ---------------- CATALOGUE ----------------
    ws_cat = wb.active
    ws_cat.title = "Catalogue"

    ws_cat.append(["Panneaux"])
    ws_cat.append(["ID", "P_STC_W", "Voc", "Vmp", "Isc", "Imp", "alpha_V_%/°C"])
    first_panel_row = ws_cat.max_row + 1
    for p in panels:
        ws_cat.append(p)
    last_panel_row = ws_cat.max_row

    ws_cat.append([""])
    ws_cat.append(["Onduleurs"])
    ws_cat.append([
        "ID", "P_AC_nom", "P_DC_max", "V_MPP_min", "V_MPP_max",
        "V_DC_max", "I_MPPT", "Nb_MPPT", "Type_reseau", "Famille"
    ])
    first_inv_row = ws_cat.max_row + 1
    for inv in inverters:
        ws_cat.append(list(inv))
    last_inv_row = ws_cat.max_row

    ws_cat.append([""])
    ws_cat.append(["Batteries"])
    ws_cat.append(["ID", "Cap_kWh"])
    first_bat_row = ws_cat.max_row + 1
    for b in batteries:
        ws_cat.append(b)
    last_bat_row = ws_cat.max_row

    _autofit(ws_cat, max_col=10)

    # ---------------- CHOIX ----------------
    ws_ch = wb.create_sheet("Choix")

    ws_ch["A1"] = "Panneau"
    ws_ch["B1"] = config.get("panel_id", "")

    ws_ch["A2"] = "Nombre modules"
    ws_ch["B2"] = config.get("n_modules", 10)

    ws_ch["A3"] = "Type réseau"
    ws_ch["B3"] = config.get("grid_type", "")

    ws_ch["A4"] = "Batterie ?"
    ws_ch["B4"] = "Oui" if config.get("battery_enabled", False) else "Non"

    ws_ch["A5"] = "Batterie (kWh)"
    ws_ch["B5"] = config.get("battery_kwh", 0.0)

    ws_ch["A6"] = "Ratio DC/AC max"
    ws_ch["B6"] = config.get("max_dc_ac", 1.5)

    ws_ch["A7"] = "Onduleur"
    ws_ch["B7"] = config.get("inverter_id", "")

    ws_ch["A9"] = "P_STC panneau"
    ws_ch["B9"] = (
        f"=IFERROR(VLOOKUP(B1,Catalogue!$A${first_panel_row}:$G${last_panel_row},2,FALSE),\"\")"
    )

    ws_ch["A10"] = "Puissance DC totale (W)"
    ws_ch["B10"] = "=IF(B9<>\"\",B9*B2,\"\")"

    _autofit(ws_ch, max_col=7)

    # ---------------- PROFIL ----------------
    ws_pr = wb.create_sheet("Profil")

    ws_pr["A1"] = "Conso annuelle (kWh)"
    ws_pr["B1"] = config.get("annual_consumption", 3500)

    ws_pr["A2"] = "Profil conso"
    ws_pr["B2"] = config.get("consumption_profile", "Standard")

    ws_pr.append([""])
    ws_pr.append(["Mois", "%_conso", "Conso_kWh", "Prod_PV_kWh", "kWh_kWp_BEL", "Autocons_kWh"])

    months = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin",
              "Juil", "Août", "Sep", "Oct", "Nov", "Déc"]
    percent_std = [7, 7, 8, 9, 9, 9, 9, 9, 8, 8, 8, 9]
    percent_winter = [10, 10, 10, 9, 8, 7, 6, 6, 7, 8, 9, 10]
    percent_summer = [6, 6, 7, 8, 9, 10, 11, 11, 10, 8, 7, 7]

    annual_kwh_kwp = 1034.0
    distribution = [3.8, 5.1, 8.7, 11.5, 12.1, 11.8, 11.9, 10.8, 9.7, 7.0, 4.3, 3.3]
    kwh_kwp = [annual_kwh_kwp * d / 100.0 for d in distribution]

    start = 5
    for i, m in enumerate(months):
        r = start + i
        ws_pr.cell(r, 1).value = m
        formula_pct = (
            '=CHOOSE(MATCH($B$2,{"Standard","Hiver fort","Été fort"},0),'
            f'{percent_std[i]},{percent_winter[i]},{percent_summer[i]})'
        )
        ws_pr.cell(r, 2).value = formula_pct
        ws_pr.cell(r, 3).value = f"=($B$1*B{r}/100)"
        ws_pr.cell(r, 5).value = kwh_kwp[i]
        ws_pr.cell(r, 4).value = f"=E{r}*Choix!$B$10/1000"
        ws_pr.cell(r, 6).value = f"=MIN(C{r},D{r})"

    _autofit(ws_pr, max_col=6)

    # ---------------- STRINGS ----------------
    ws_st = wb.create_sheet("Strings")

    ws_st["A1"] = "Vérification string"
    ws_st["A3"] = "Panneau"
    ws_st["B3"] = config.get("panel_id", "")
    ws_st["A4"] = "Onduleur"
    ws_st["B4"] = config.get("inverter_id", "")

    ws_st["A5"] = "T° min"
    ws_st["B5"] = config.get("t_min", -10)

    ws_st["A6"] = "T° max"
    ws_st["B6"] = config.get("t_max", 70)

    ws_st["A7"] = "Modules en série"
    ws_st["B7"] = config.get("n_series", 10)

    ws_st["A9"] = "Voc module"
    ws_st["B9"] = (
        f"=IFERROR(VLOOKUP(B3,Catalogue!$A${first_panel_row}:$G${last_panel_row},3,FALSE),\"\")"
    )

    ws_st["A10"] = "Vmp module"
    ws_st["B10"] = (
        f"=IFERROR(VLOOKUP(B3,Catalogue!$A${first_panel_row}:$G${last_panel_row},4,FALSE),\"\")"
    )

    ws_st["A11"] = "α_V (%/°C)"
    ws_st["B11"] = (
        f"=IFERROR(VLOOKUP(B3,Catalogue!$A${first_panel_row}:$G${last_panel_row},7,FALSE),\"\")"
    )

    ws_st["A13"] = "V_DC_max"
    ws_st["B13"] = (
        f"=IFERROR(VLOOKUP(B4,Catalogue!$A${first_inv_row}:$J${last_inv_row},6,FALSE),\"\")"
    )

    ws_st["A14"] = "V_MPP_min"
    ws_st["B14"] = (
        f"=IFERROR(VLOOKUP(B4,Catalogue!$A${first_inv_row}:$J${last_inv_row},4,FALSE),\"\")"
    )

    ws_st["A15"] = "V_MPP_max"
    ws_st["B15"] = (
        f"=IFERROR(VLOOKUP(B4,Catalogue!$A${first_inv_row}:$J${last_inv_row},5,FALSE),\"\")"
    )

    ws_st["A17"] = "Voc string froid"
    ws_st["B17"] = "=B7*B9*(1+B11/100*(B5-25))"

    ws_st["A18"] = "Vmp string chaud"
    ws_st["B18"] = "=B7*B10*(1+B11/100*(B6-25))"

    ws_st["A20"] = "Check Voc <= V_DC_max"
    ws_st["B20"] = "=IF(AND(B17<>\"\",B13<>\"\"),IF(B17<=B13,\"OK\",\"DÉPASSE\"),\"\")"

    ws_st["A21"] = "Check Vmp dans MPPT"
    ws_st["B21"] = "=IF(AND(B18<>\"\",B14<>\"\",B15<>\"\"),IF(AND(B18>=B14,B18<=B15),\"OK\",\"HORS PLAGE\"),\"\")"

    _autofit(ws_st, max_col=4)

    # ---------------- SYNTHÈSE ----------------
    ws_sy = wb.create_sheet("Synthese")

    ws_sy["A1"] = "Synthèse client"

    ws_sy["A3"] = "Panneau"
    ws_sy["B3"] = "=Choix!B1"

    ws_sy["A4"] = "Modules"
    ws_sy["B4"] = "=Choix!B2"

    ws_sy["A5"] = "Puissance DC totale"
    ws_sy["B5"] = "=Choix!B10"

    ws_sy["A7"] = "Onduleur"
    ws_sy["B7"] = "=Choix!B7"

    ws_sy["A9"] = "Conso annuelle"
    ws_sy["B9"] = "=Profil!B1"

    ws_sy["A10"] = "Prod PV annuelle"
    ws_sy["B10"] = "=SUM(Profil!D5:D16)"

    ws_sy["A11"] = "Autocons annuelle"
    ws_sy["B11"] = "=SUM(Profil!F5:F16)"

    ws_sy["A12"] = "Taux autocons"
    ws_sy["B12"] = "=IF(B10>0,B11/B10,\"\")"

    ws_sy["A13"] = "Taux couverture"
    ws_sy["B13"] = "=IF(B9>0,B11/B9,\"\")"

    ws_sy["A15"] = "Batterie ?"
    ws_sy["B15"] = "=Choix!B4"

    ws_sy["A16"] = "Capacité batterie (kWh)"
    ws_sy["B16"] = "=Choix!B5"

    ws_sy["A17"] = "Modèle batterie"
    ws_sy["B17"] = "=IF(B15<>\"Oui\",\"Aucune\",IF(B16<=6,\"Sigen6\",\"Sigen10\"))"

    _autofit(ws_sy, max_col=4)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
